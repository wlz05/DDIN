# DDIN: Domain-aware Disentangled Interaction Network for Multimodal Fake News Detection

import openpyxl
    Deletion rules:
    1. Keyword filter,
    2. Aspect ratio severely imbalanced or too small,
    3. Specified image sets,
    4. Must be openable by cv2,
    5. Must not be questions,
    6. Text deduplication,
    7. Image deduplication
forbidden_words = [
    "zombie", "ring", "sleep_with", "bed",
    "gang_rape", "mistress", "rape", "too_cruel", "steal_child", "idiot", "dumb", "moron", "retard",
    "moron2", "damn", "prostitute", "wtf", "fuck", "motherfucker", "bastard", "son_of_bitch", "crematorium"
]
absolute_fake_words = ["missing_person", "urgent_find_child"]
not_allowed_words = forbidden_words + absolute_fake_words
thresh_width_height = 2
not_allowed_clusters = ['./hashing_self_repeat/343']
baned_hashing = set()

rumor_root = './data/rumor_images/'
weibo_fake_root = './data/rumor_images/'
weibo_real_root = './data/nonrumor_images/'
weibo21_fake_root = './weibo21/rumor_images/'
weibo21_real_root = './weibo21/nonrumor_images/'


mixset_xlsx = "./dataset/rumor_dataset/all_images.xlsx"

def get_workbook(xlsx):
    wb = openpyxl.load_workbook(xlsx)
    sheetnames = wb.sheetnames
    sheet = wb[sheetnames[0]]
    rows = sheet.max_row
    return sheet, rows


def strQ2B(ustring):
    rstring = ""
    for uchar in ustring:
        inside_code = ord(uchar)
        if inside_code == 12288:  # full-width space
            inside_code = 32
        elif 65281 <= inside_code <= 65374:  # full-width chars (except space)
            inside_code -= 65248
        if inside_code <= 66812:
            rstring += chr(inside_code)
    return rstring

def del_emoji(ustring):
    rstring = ""
    for uchar in ustring:
        inside_code = ord(uchar)
        if inside_code <= 66812:
            rstring += chr(inside_code)
    return rstring
